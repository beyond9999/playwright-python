import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

response = requests.get('https://www.momoyu.cc/api/hot/list?type=0')

if response.status_code == 200:
    data = response.json()['data']
    wb = Workbook()
    wb.remove(wb['Sheet'])

    for i in data:
        name = i['name']
        ws = wb.create_sheet(title=name)

        for entry in i['data']:
            title = entry['title']
            link = entry['link']
            ws.append([title])

            title_cell = ws.cell(row=ws.max_row, column=1)
            hyperlink = Hyperlink(ref=f"{link}", target=link)
            title_cell.hyperlink = hyperlink
            title_cell.style = "Hyperlink"
            title_cell.font = Font(size=11)

        ws.column_dimensions['A'].width = 100

    wb.save("momoyu.xlsx")
    print("数据已存储到文件中")
else:
    print('请求失败:', response.status_code)
